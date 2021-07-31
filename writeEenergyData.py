from openpyxl import load_workbook
 
#load excel file
workbook = load_workbook(filename="edat.xlsx")
sheet_name="A1"
row_max=1000
 
#open workbook
sheet = workbook.active

#open text file to read
file1 = open('vzr.txt', 'r')
count = 0
 
#loop over text file line by line

for line in file1:
#parse data in the line and read into two variables
    varList=line.split()
    varList[0]=int(varList[0])
    varList[1]=float(varList[1])
    #print("varList[0]:", varList[0],"\n")
    #print(varList[1])
	#loop over excel sheet column, read first colume in a variable
    for i in range(1, row_max+1):
        #print("Row ", i, " data :")
        #for j in range(1, sheet.max_column+1):
        cell_obj = sheet.cell(row=i, column=1)
        energy_cell=sheet.cell(row=i,column=7)
        #print(cell_obj.value,"\n ")	
	    # if match
        if (cell_obj.value == varList[0]):
            print("matched for structure index: ",varList[0])
            energy_cell.value=varList[1] #modify the desired cell
            # sheet["A1"] = "Full Name"
            # count += 1
    #print("Line{}: {}".format(count, line.strip()))
 
# Closing files
file1.close()

#save the file
workbook.save(filename="edat.xlsx")

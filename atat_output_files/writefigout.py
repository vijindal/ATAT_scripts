import os
from openpyxl import load_workbook

#path = os.getcwd()
#print(path)
 
#load excel file
fileName="atat_output_files/figout.xlsx"
workbook = load_workbook(filename=fileName)

#updating 'clusinfo.out' sheet
#  open workbook 
sheet_name="clusinfo.out"
sheet = workbook[sheet_name]
#print(sheet.title)
#open clusinfo.out' to read
file1 = open('atat_output_files/clusinfo.out', 'r')
#loop over text file line by line
line_number=1
for line in file1:
    #print("line number:", line_number)
#parse data in the line and read into four variables (num of sites,	distance, multiplicity,	eci)
    varList=line.split()
    varList[0]=int(varList[0]) #num of sites
    varList[1]=float(varList[1]) #distance
    varList[2]=int(varList[2]) #multiplicity
    varList[3]=float(varList[3]) #eci
    #print("varList[0]:", varList[0])
    #print("varList[1]:", varList[1])
    #print("varList[2]:", varList[2])
    #print("varList[3]:", varList[3])
	#write data in excel sheet column
    numsite_cell=sheet.cell(row=line_number+10,column=1)
    distance_cell=sheet.cell(row=line_number+10,column=2)
    multiplicity_cell=sheet.cell(row=line_number+10,column=3)
    eci_cell=sheet.cell(row=line_number+10,column=4)
    numsite_cell.value=varList[0]
    distance_cell.value=varList[1]
    multiplicity_cell.value=varList[2]
    eci_cell.value=varList[3]
    line_number=line_number+1
#open 'eci.out' to read
file1 = open('atat_output_files/eci.out', 'r')
#loop over text file line by line
line_number=1
for line in file1:
    #parse data in the line and read into variables (eci)
    varList=line.split()
    varList[0]=float(varList[0]) #eci
    #print("varList[0]:", varList[0])
    eci_cell=sheet.cell(row=line_number+8,column=6)
    eci_cell.value=varList[0]
    line_number=line_number+1

#updating 'fit.out' sheet
#  open workbook 
sheet_name="fit.out"
sheet = workbook[sheet_name]
#print(sheet.title)
#open 'fit.out' to read
file1 = open('atat_output_files/fit.out', 'r')
#loop over text file line by line
line_number=2
for line in file1:
#parse data in the line and read into six variables:concentration,	energy,	fitted_energy,	(energy-fitted_energy),	weight, 	index 
    varList=line.split()
    varList[0]=float(varList[0]) #concentration
    varList[1]=float(varList[1]) #energy
    varList[2]=float(varList[2]) #fitted_energy
    varList[3]=float(varList[3]) #(energy-fitted_energy)
    varList[4]=float(varList[4]) #weight
    varList[5]=int(varList[5]) #index
  
	#write data in excel sheet column
    data0_cell=sheet.cell(row=line_number,column=1)
    data1_cell=sheet.cell(row=line_number,column=2)
    data2_cell=sheet.cell(row=line_number,column=3)
    data3_cell=sheet.cell(row=line_number,column=4)
    data4_cell=sheet.cell(row=line_number,column=5)
    data5_cell=sheet.cell(row=line_number,column=6)
    data0_cell.value=varList[0]
    data1_cell.value=varList[1]
    data2_cell.value=varList[2]
    data3_cell.value=varList[3]
    data4_cell.value=varList[4]
    data5_cell.value=varList[5]
    line_number=line_number+1

#updating 'gs.out' sheet
#  open workbook 
sheet_name="gs.out"
sheet = workbook[sheet_name]
#print(sheet.title)
#open 'gs.out' to read
file1 = open('atat_output_files/gs.out', 'r')
#loop over text file line by line
line_number=2
for line in file1:
#parse data in the line and read into four variables:concentration	energy	fitted_energy	index
    varList=line.split()
    varList[0]=float(varList[0]) #concentration
    varList[1]=float(varList[1]) #energy
    varList[2]=float(varList[2]) #fitted_energy
    varList[3]=int(varList[3]) #index
	#write data in excel sheet column
    data0_cell=sheet.cell(row=line_number,column=1)
    data1_cell=sheet.cell(row=line_number,column=2)
    data2_cell=sheet.cell(row=line_number,column=3)
    data3_cell=sheet.cell(row=line_number,column=4)
    data0_cell.value=varList[0]
    data1_cell.value=varList[1]
    data2_cell.value=varList[2]
    data3_cell.value=varList[3]
    print("varList[0]:", varList[0],data0_cell.value)
    print("varList[1]:", varList[1],data1_cell.value)
    print("varList[2]:", varList[2],data2_cell.value)
    print("varList[3]:", varList[3],data3_cell.value)
    line_number=line_number+1

# Closing files
file1.close()

#save the file
workbook.save(filename=fileName)
